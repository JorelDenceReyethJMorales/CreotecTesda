from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import os
from werkzeug.utils import secure_filename
from app.routes.auth import auth_bp
from app.routes.generate import bp as generate_bp
from app.routes.upload import bp as upload_bp

# --- extra imports for /api/generate ---
import io, json, re, traceback
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# Define where templates are stored (your existing code)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads", "templates")
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Register blueprints (your existing code)
app.register_blueprint(auth_bp)
app.register_blueprint(generate_bp)
app.register_blueprint(upload_bp)

DEFAULT_MAPPING = {}

PLACEHOLDER_RE = re.compile(r"\{([^}]+)\}")

def format_value(val, fmt=None):
    return "" if val is None else str(val)

def replace_placeholders_in_cell(text, mapping, rowdict):
    # Handle YEAR LAST ATTENDED context switching as in your code
    if "YEAR LAST ATTENDED" in text.upper():
        context = None
        up = text.upper()
        if "ELEMENTARY" in up:
            context = "ELEMENTARY"
        elif "SECONDARY" in up:
            context = "SECONDARY"
        elif "TERTIARY" in up:
            context = "TERTIARY"
    else:
        context = None

    def repl(m):
        key = m.group(1)
        mp = mapping.get(key, key)
        if isinstance(mp, dict):
            col = mp.get(context) or mp.get("DEFAULT")
        else:
            col = mp
        val = rowdict.get(col, "")
        return format_value(val)

    return PLACEHOLDER_RE.sub(repl, text)

def replace_placeholders_in_worksheet(ws, mapping, rowdict):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and "{" in cell.value and "}" in cell.value:
                cell.value = replace_placeholders_in_cell(cell.value, mapping, rowdict)

def _safe_sheet_title(s: str, used: set) -> str:
    title = (s or "").strip() or "Row"
    for ch in '[]:*?/\\':
        title = title.replace(ch, "-")
    title = title[:31] or "Row"
    orig = title
    i = 2
    while title in used:
        suffix = f" ({i})"
        title = (orig[: 31 - len(suffix)] + suffix) if len(orig) + len(suffix) > 31 else orig + suffix
        i += 1
    used.add(title)
    return title

def _copy_template_sheet_with_fallback(wb, template_ws, new_title):
    try:
        ws_copy = wb.copy_worksheet(template_ws)
        ws_copy.title = new_title
        return ws_copy
    except Exception as e:
        print("[WARN] copy_worksheet failed; falling back to manual copy:", repr(e))
        ws = wb.create_sheet(title=new_title)
        for rng in template_ws.merged_cells.ranges:
            ws.merge_cells(str(rng))
        max_row = template_ws.max_row
        max_col = template_ws.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = template_ws.cell(row=r, column=c).value
                if v is not None:
                    ws.cell(row=r, column=c, value=v)
        return ws

@app.route("/api/generate", methods=["POST"])
def generate_excel():
    """
    Upload a user Excel file with at least two sheets (details, grades),
    fill the server-side template.xlsx with placeholders, and return the
    generated workbook (one sheet per row).
    """
    def err(msg, status=400, exc=None):
        print(f"[ERROR] {msg}")
        if exc:
            traceback.print_exc()
        return jsonify({"error": msg}), status

    # 1) Validate upload
    if "file" not in request.files:
        return err("No file part 'file'")
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return err("Please upload an .xlsx or .xlsm file")

    # 2) Optional mapping JSON
    mapping = DEFAULT_MAPPING.copy()
    if "mapping" in request.form and request.form["mapping"].strip():
        try:
            mapping.update(json.loads(request.form["mapping"]))
        except Exception as e:
            return err(f"Invalid mapping JSON: {e}", exc=e)

    # 3) Read uploaded workbook (all sheets)
    try:
        xl = pd.read_excel(f, sheet_name=None, dtype=str)
        xl = {k: v.fillna("") for k, v in xl.items()}
    except Exception as e:
        return err(f"Could not read Excel: {e}", exc=e)

    if len(xl) < 2:
        return err("Uploaded file must have at least 2 worksheets: details and grades.")

    sheet_names = list(xl.keys())
    df_details = xl[sheet_names[0]]
    df_grades = xl[sheet_names[1]]

    if df_details.empty:
        return err("Details sheet has no rows.")

    # 4) Load the template
    if not os.path.exists(UPLOAD_FOLDER + "/template.xlsx"):
        return err("Template not found on server", status=500)
    try:
        base_wb = load_workbook(UPLOAD_FOLDER + "/template.xlsx", data_only=True)
    except Exception as e:
        return err(f"Could not open template: {e}", status=500, exc=e)

    if not base_wb.worksheets:
        return err("Template has no worksheets.", status=500)
    template_ws = base_wb.worksheets[0]

    # 5) Fill per row
    used_titles = set()
    for idx, row in df_details.iterrows():
        row_dict = row.to_dict()

        # Find the key mapped to NAME (default "NAME")
        name_key = next((k for k in mapping.keys() if k.upper() == "NAME"), "NAME")
        col_for_name = mapping.get(name_key, name_key)
        candidate_name = row_dict.get(col_for_name, f"Row {idx+1}")
        new_title = _safe_sheet_title(str(candidate_name), used_titles)

        # Copy template sheet and compute combined row (details + first matched grades)
        ws_copy = _copy_template_sheet_with_fallback(base_wb, template_ws, new_title)
        matched_grades = df_grades[df_grades[col_for_name] == candidate_name]
        grade_row = matched_grades.iloc[0].to_dict() if not matched_grades.empty else {}
        combined_row = {**row_dict, **grade_row}

        try:
            replace_placeholders_in_worksheet(ws_copy, mapping, combined_row)
        except Exception as e:
            return err(f"Failed to fill placeholders on sheet '{new_title}': {e}", status=500, exc=e)

    # Remove the original template sheet so the output contains only generated sheets
    base_wb.remove(template_ws)

    # 6) Return the workbook as a download
    try:
        output = io.BytesIO()
        base_wb.save(output)
        output.seek(0)
    except Exception as e:
        return err(f"Failed to save generated workbook: {e}", status=500, exc=e)

    out_name = f"filled_multi_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=out_name,
    )

# ======================================================
# END: Excel -> Excel generator
# ======================================================

# ---- your existing endpoints below (unchanged) ----

@app.route('/')
def home():
    return "Hello, Creo Certificate Backend!"

# Upload template endpoint (existing)
@app.route("/api/upload-template", methods=["POST"])
def upload_template():
    file = request.files.get("template")
    template_type = request.form.get("type")  # ojt, immersion, custom

    if not file or not template_type:
        return jsonify({"error": "Missing file or type"}), 400

    if not file.filename.endswith((".ppt", ".pptx")):
        return jsonify({"error": "Only .ppt and .pptx files allowed"}), 400

    # Save the file with a fixed name: ojt.pptx, immersion.pptx, or custom.pptx
    filename = f"{template_type}.pptx"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(save_path)

    return jsonify({"message": "Upload successful", "filename": filename}), 200

# List available templates (existing)
@app.route("/api/templates", methods=["GET"])
def list_templates():
    data = {}

    # Default files (only for ojt and immersion)
    default_files = {
        "ojt": "ojt_default.pptx",
        "immersion": "immersion_default.pptx"
    }

    for key, filename in default_files.items():
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(path):
            data[key] = {
                "name": filename,
                "isDefault": True,
                "url": f"/uploads/templates/{filename}"
            }

    # Uploaded custom template (only if user uploaded it)
    custom_path = os.path.join(app.config['UPLOAD_FOLDER'], "custom.pptx")
    if os.path.exists(custom_path):
        data["custom"] = {
            "name": "custom.pptx",
            "isDefault": False,
            "url": "/uploads/templates/custom.pptx"
        }

    return jsonify(data)

# Serve a template file for download (existing)
@app.route("/uploads/templates/<filename>")
def serve_template(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == "__main__":
    # Keep your existing bind/port
    app.run(debug=True, host='0.0.0.0', port=5000)
